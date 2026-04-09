#!/usr/bin/env python3
"""
Universal Excel Builder – JSON Driven
=====================================
Usage:
    python3 universal_builder.py path/to/export_data.json

Expects JSON Schema:
{
  "output_path": "docs/Report.xlsx",
  "title": "BÁO CÁO",
  "subtitle": "Ngày xuất báo cáo",
  "palette": {
      "header_bg": "1F4E79", "header_ft": "FFFFFF",
      "row_even": "DEEAF1", "row_odd": "FFFFFF",
      "border": "BDD7EE", "sect_bg": "2E75B6", "sect_ft": "FFFFFF"
  },
  "color_mapping": { 
      "Lỗi": "F4CCCC", "Cảnh báo": "FFF2CC", "Thêm mới": "D9EAD3" 
  },
  "legend": [
      {"label": "Lỗi (Error)", "color": "F4CCCC"},
      {"label": "Cảnh báo (Warning)", "color": "FFF2CC"}
  ],
  "summary_sheet": { 
      "name": "TỔNG HỢP", 
      "headers": [{"label": "Hạng mục", "width": 20}, {"label": "Số lượng", "width": 10}],
      "add_total_row": true,
      "rows": [ ["Module A", 5] ] 
  },
  "detail_sheets": [
      {
          "name": "Chi tiết Module A",
          "color_col_index": 2, 
          "headers": [
              {"label": "STT", "width": 5}, 
              {"label": "Mô tả", "width": 40}, 
              {"label": "Loại", "width": 15}
          ],
          "rows": [ [1, "Thiếu nút Save", "Lỗi"] ]
      }
  ]
}
"""
import os
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 universal_builder.py <path_to_json>")
        sys.exit(1)

    json_path = sys.argv[1]
    if not os.path.exists(json_path):
        print(f"File not found: {json_path}")
        sys.exit(1)

    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    # Defaults Palette
    palette = config.get("palette", {})
    C_HEADER_BG = palette.get("header_bg", "1F4E79")
    C_HEADER_FT = palette.get("header_ft", "FFFFFF")
    C_SECT_BG   = palette.get("sect_bg", "2E75B6")
    C_SECT_FT   = palette.get("sect_ft", "FFFFFF")
    C_ALT1      = palette.get("row_even", "DEEAF1")
    C_ALT2      = palette.get("row_odd", "FFFFFF")
    C_BORDER    = palette.get("border", "BDD7EE")
    TYPE_COLOR  = config.get("color_mapping", {})

    def hfill(hx):
        return PatternFill("solid", fgColor=hx.replace("#", ""))

    def tborder():
        s = Side(style="thin", color=C_BORDER.replace("#", ""))
        return Border(left=s, right=s, top=s, bottom=s)

    def row_height(values, headers_widths):
        max_height = 15
        for i, v in enumerate(values):
            col_width = headers_widths[i] if i < len(headers_widths) else 38
            text_len = len(str(v))
            lines = text_len // col_width + str(v).count("\n") + 1
            max_height = max(max_height, lines * 14)
        return min(max_height, 120)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. SUMMARY SHEET
    sum_cfg = config.get("summary_sheet")
    if sum_cfg:
        sh_name = sum_cfg.get("name", "TỔNG HỢP")
        ws = wb.create_sheet(sh_name, 0)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        headers = sum_cfg.get("headers", [])
        ncols = len(headers)
        if ncols == 0: ncols = 1

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=config.get("title", "SUMMARY"))
        c.font = Font(bold=True, size=13, color=C_HEADER_FT, name="Calibri")
        c.fill = hfill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=config.get("subtitle", ""))
        c.font = Font(italic=True, size=9, color="595959", name="Calibri")
        c.fill = hfill("F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 14

        # Header Row
        ws.row_dimensions[3].height = 22
        widths = []
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h.get("label", ""))
            c.font = Font(bold=True, color=C_HEADER_FT, size=10, name="Calibri")
            c.fill = hfill(C_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = tborder()
            w = h.get("width", 15)
            widths.append(w)
            ws.column_dimensions[get_column_letter(i)].width = w

        # Data Rows
        rows = sum_cfg.get("rows", [])
        for ri, row_vals in enumerate(rows, 4):
            bg = C_ALT1 if ri % 2 == 0 else C_ALT2
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = hfill(bg)
                c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = tborder()
            ws.row_dimensions[ri].height = 20

        # Total Row
        total_row_idx = len(rows) + 4
        if sum_cfg.get("add_total_row", False) and len(rows) > 0:
            ws.row_dimensions[total_row_idx].height = 20
            # Calculate sum for numeric columns
            totals = ["TỔNG"] + [
                sum(r[i] for r in rows if isinstance(r[i], (int, float)))
                for i in range(1, ncols)
            ]
            for ci, v in enumerate(totals, 1):
                c = ws.cell(row=total_row_idx, column=ci, value=v if v != 0 else "")
                if ci == 1: c.value = "TỔNG"
                c.fill = hfill(C_SECT_BG)
                c.font = Font(bold=True, size=11, color=C_SECT_FT, name="Calibri")
                c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
                c.border = tborder()
            total_row_idx += 1

        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{len(rows)+3}"

        # Legend
        legend = config.get("legend", [])
        if legend:
            start_row = total_row_idx + 1
            ws.merge_cells(f"A{start_row}:{get_column_letter(ncols)}{start_row}")
            lh = ws.cell(row=start_row, column=1, value="CHÚ GIẢI MÀU SẮC:")
            lh.font = Font(bold=True, size=10, name="Calibri")
            ws.row_dimensions[start_row].height = 16
            for i, leg in enumerate(legend, start_row + 1):
                ws.row_dimensions[i].height = 14
                ws.merge_cells(f"A{i}:{get_column_letter(ncols)}{i}")
                lc = ws.cell(row=i, column=1, value=f"  ■  {leg.get('label')}")
                lc.fill = hfill(leg.get("color", "FFFFFF"))
                lc.font = Font(size=9, name="Calibri")
                lc.border = tborder()

    # 2. DETAIL SHEETS
    for d_idx, d_cfg in enumerate(config.get("detail_sheets", [])):
        sh_name = d_cfg.get("name", f"Sheet {d_idx+1}")
        # Truncate sheet name up to 31 chars
        ws = wb.create_sheet(sh_name[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

        headers = d_cfg.get("headers", [])
        ncols = len(headers)
        if ncols == 0: ncols = 1
        color_idx = d_cfg.get("color_col_index", -1)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=d_cfg.get("title", config.get("title", sh_name)))
        c.font = Font(bold=True, size=13, color=C_HEADER_FT, name="Calibri")
        c.fill = hfill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Header
        ws.row_dimensions[2].height = 22
        widths = []
        for i, h in enumerate(headers, 1):
            ws.cell(row=2, column=i, value=h.get("label", ""))
            c = ws.cell(row=2, column=i)
            c.font = Font(bold=True, color=C_HEADER_FT, size=10, name="Calibri")
            c.fill = hfill(C_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = tborder()
            w = h.get("width", 15)
            widths.append(w)
            ws.column_dimensions[get_column_letter(i)].width = w

        # Rows
        drows = d_cfg.get("rows", [])
        for ri, row_vals in enumerate(drows, 3):
            # Evaluate type color based on 0-indexed color_col_index
            row_color = ""
            if len(row_vals) > color_idx and color_idx >= 0:
                row_type = row_vals[color_idx]
                row_color = TYPE_COLOR.get(str(row_type), "")
            
            # Alternative: config default "color_col_val" force for entire sheet (like Open Q)
            if d_cfg.get("force_color"):
                row_color = d_cfg.get("force_color")

            bg = row_color.replace("#", "") if row_color else (C_ALT1 if ri % 2 == 0 else C_ALT2)
            
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = hfill(bg)
                c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = tborder()
            ws.row_dimensions[ri].height = row_height(row_vals, widths)

        ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(drows)+2}"


    # FALLBACK if no sheets
    if len(wb.worksheets) == 0:
        wb.create_sheet("Empty Report")

    # SAVE
    out_path = config.get("output_path", "docs/Export.xlsx")
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)

    wb.save(out_path)
    print(f"✅ Báo cáo Excel đã được xuất thành công: {out_path}")

if __name__ == "__main__":
    main()
